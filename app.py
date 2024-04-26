import itertools
import pandas as pd
import re
from itertools import chain
from fuzzywuzzy import fuzz
from openpyxl.styles import PatternFill

def standardize_phone_numbers(phone_numbers):
    """
    Standardize phone numbers by removing non-digit characters, 
    keeping "MISSING" values unchanged.
    """
    return {num if num.strip().upper() == 'MISSING' else re.sub(r'\D', '', num) 
            for num in phone_numbers.split('|') if num}
    # return {re.sub(r'\D', '', num) for num in phone_numbers.split('|') if num and num.strip().upper() != 'MISSING'}


def standardize_emails(emails):
    """
    Standardize emails by converting to lowercase and removing spaces.
    keeping "MISSING" values unchanged.
    """
    return {email if email.strip().upper() == 'MISSING' else email.strip().lower() for email in emails.split('|') if email}
    # return {email.strip().lower() for email in emails.split('|') if email and email.strip().upper() != 'MISSING'}

def full_match_in_cluster(all_values): # correct
    """
    Check if all values in a cluster are exactly the same across all rows,
    excluding certain keywords that invalidate a match.
    """
    excluded_values = {'MISSING', 'missing'}
    consolidated_values = set(chain.from_iterable(all_values))
    
    # Check if there's only one unique value that isn't in the excluded set
    if len(consolidated_values) == 1 and not consolidated_values.intersection(excluded_values):
        return True
    return False


# def linked_match_in_cluster(all_values): # bad. use the version below
#     """
#     Check if there is a linked (indirect or chained) match within the cluster.
#     """
#     linked = set()
#     for values in all_values:
#         if not linked:
#             linked.update(values)
#         else:
#             linked.intersection_update(values)
#             if not linked:
#                 return False
#     return True

def linked_match_in_cluster(all_values): # should be correct
    """
    Check if there is a linked (indirect or chained) match within the cluster,
    excluding sets containing only "MISSING".

    Checks if the rows in the cluster are same person by taking all values from a single column and comparing them
    """
    linked = set()
    # print(all_values)
    for i, values in enumerate(all_values):
        if values == {'MISSING'} or values == {'missing'}:
            # continue  # Skip sets with only "MISSING" or 'missing'
            return False
        # if not linked:
        #     linked.update(values)
        if i == 0:
            linked.update(values)
        else:
            linked.intersection_update(values)
            # if not linked:
            #     return False
    
    # Check if any links were established 
    return bool(linked)
    # return bool(linked) and len(linked) == 1  # Return True if linked is not empty, False otherwise

def process_clusters(data):
    # How many clusters were classified successfully in each pass
    first_pass_match_count = 0
    second_pass_match_count = 0
    third_pass_match_count = 0
    fourth_pass_match_count = 0
    # First Pass: Direct match in each cluster
    for investigator_id, cluster in data.groupby('fire_investigator_id'):
        all_names = cluster['investigator_full_name'].apply(lambda x: {x.strip().lower()})
        all_phones = cluster['investigator_phone_number'].apply(standardize_phone_numbers)
        all_emails = cluster['investigator_email'].apply(standardize_emails)

        name_match = full_match_in_cluster(all_names)
        phone_match = full_match_in_cluster(all_phones)
        email_match = full_match_in_cluster(all_emails)

        # if name_match:
        #     data.loc[cluster.index, 'same'] = 'first pass'
        #     data.loc[cluster.index, 'matching_criteria'] = 'all names match'
        
        # elif phone_match:
        if phone_match or email_match:
            data.loc[cluster.index, 'same'] = 'first pass'
            first_pass_match_count += 1
            if phone_match:
                data.loc[cluster.index, 'matching_criteria'] = 'all phone numbers match'
            elif email_match:
                data.loc[cluster.index, 'matching_criteria'] = 'all emails match'
    print(f"first_pass_match_count: {first_pass_match_count}")

    # Second Pass: Linked match in each cluster
    for investigator_id, cluster in data.groupby('fire_investigator_id'):
        if data.loc[cluster.index, 'same'].iloc[0] == 'first pass':
            continue  # Skip this cluster as it has already matched in the first pass

        all_names = cluster['investigator_full_name'].apply(lambda x: {x.strip().lower()})
        all_phones = cluster['investigator_phone_number'].apply(standardize_phone_numbers)
        all_emails = cluster['investigator_email'].apply(standardize_emails)

        name_linked = linked_match_in_cluster(all_names)
        phone_linked = linked_match_in_cluster(all_phones)
        email_linked = linked_match_in_cluster(all_emails)

        if phone_linked or email_linked:
            second_pass_match_count += 1
            data.loc[cluster.index, 'same'] = 'second pass'
            matching_criteria = []
            if phone_linked:
                matching_criteria.append('phone numbers')
            if email_linked:
                matching_criteria.append('emails')
            data.loc[cluster.index, 'matching_criteria'] = ' '.join(matching_criteria)
    print(f"second_pass_match_count: {second_pass_match_count}")

    # Third Pass: Cross-Link Matching in each cluster
    for investigator_id, cluster in data.groupby('fire_investigator_id'):
        if data.loc[cluster.index, 'same'].notnull().all():
            continue  # Skip this cluster if all rows are already matched in previous passes
        
        # Standardize phone numbers and emails before comparison
        cluster['standardized_phone'] = cluster['investigator_phone_number'].apply(standardize_phone_numbers)
        cluster['standardized_email'] = cluster['investigator_email'].apply(standardize_emails)

        connected_components = []
        row_matches = {index: [] for index in cluster.index}  # Prepare to track matches per row

        for index, row in cluster.iterrows():
            for j, compare_row in cluster.iterrows():
                if index < j:  # Ensure each pair is only considered once to avoid duplicate entries
                    matched_on = []

                    phones_intersection = row['standardized_phone'].intersection(compare_row['standardized_phone'])
                    emails_intersection = row['standardized_email'].intersection(compare_row['standardized_email'])

                    # added corrective measure in third pass:
                    # check that the intersection is not equal to "MISSING" or "missing"
                    if phones_intersection and not phones_intersection == {'MISSING'} and not phones_intersection == {'missing'}:
                        matched_on.append('phone')
                    if emails_intersection and not emails_intersection == {'MISSING'} and not emails_intersection == {'missing'}:
                        matched_on.append('email')


                    if matched_on:
                        matched_indices = {index, j}
                        match_string = f"{j+2} - {', '.join(matched_on)}"  # Creating match string
                        row_matches[index].append(match_string)
                        match_string = f"{index+2} - {', '.join(matched_on)}"  # Creating match string for the other row
                        row_matches[j].append(match_string)

                        # Union with any intersecting existing sets
                        new_components = []
                        added = False
                        for component in connected_components:
                            if not matched_indices.isdisjoint(component):
                                component.update(matched_indices)
                                added = True
                                break  # Prevent multiple additions
                        if not added:
                            new_components.append(matched_indices)
                        connected_components += new_components

        # Check if union of all components covers the entire cluster
        # BUG below
        # union_of_matches = set()
        # for component in connected_components:
        #     union_of_matches.update(component)

        # if len(union_of_matches) == len(cluster):
        # Check if there's only one connected component and it contains all rows 
        if len(connected_components) == 1 and len(connected_components[0]) == len(cluster):
            third_pass_match_count += 1
            data.loc[cluster.index, 'same'] = 'third pass'
            # Assign matches information for each row specifically
            for idx in cluster.index:
                if row_matches[idx]:
                    matches_str = "; ".join(row_matches[idx])
                    data.loc[idx, 'matching_criteria'] = f'matched with rows: {matches_str}'
                else:
                    data.loc[idx, 'matching_criteria'] = 'matched with rows: None'
    print(f"third_pass_match_count: {third_pass_match_count}")

    # Fourth Pass: Exact Match on Names and Location
    for investigator_id, cluster in data.groupby('fire_investigator_id'):
        if data.loc[cluster.index, 'same'].notnull().all():
            continue  # Skip if already matched

        all_names = cluster['investigator_full_name'].apply(lambda x: {x.strip().lower()}) 
        all_countries = cluster['investigator_country'].apply(lambda x: {x.strip().lower()})
        all_states = cluster['investigator_state'].apply(lambda x: {x.strip().lower()})
        all_cities = cluster['investigator_city'].apply(lambda x: {x.strip().lower()})

        if full_match_in_cluster(all_names) and full_match_in_cluster(all_countries) and full_match_in_cluster(all_states) and full_match_in_cluster(all_cities):
            fourth_pass_match_count += 1
            data.loc[cluster.index, 'same'] = 'fourth pass'
            data.loc[cluster.index, 'matching_criteria'] = 'Names, country, state, city are exactly matching'
    print(f"fourth_pass_match_count: {fourth_pass_match_count}")
    total_clusters = data['fire_investigator_id'].nunique()
    total_matches = first_pass_match_count+second_pass_match_count+third_pass_match_count+fourth_pass_match_count
    # print(f"Total match count: {first_pass_match_count}+{second_pass_match_count}+{third_pass_match_count}+{fourth_pass_match_count} = {total_matches}")
    print(f"Total match count: {total_matches}")
    print(f'Total number of clusters: {total_clusters}')
    print(f"Percentage of clusters found to have one and the same person: {total_matches}/{total_clusters} = {total_matches*100/total_clusters:.3f}%")

def main():
    input_file = 'input.xlsx'
    output_file = 'output.xlsx'
    data = pd.read_excel(input_file, dtype=str).fillna('MISSING')

    process_clusters(data)

    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    data.to_excel(writer, index=False, sheet_name='Results')
    workbook = writer.book
    worksheet = writer.sheets['Results']

    # Apply different colors based on match type
    for idx, row in data.iterrows():
        row_num = idx + 2  # +2 to account for header row
        if row['same'] == 'first pass':
            fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Green for first pass
        elif row['same'] == 'second pass':
            fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Light green for second pass
        elif row['same'] == 'third pass':
            fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')  # Blue for third pass
        elif row['same'] == 'fourth pass':
            fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Yellow for fourth pass
        else:
            fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White for unmatched rows


        for cell in worksheet[row_num]:
            cell.fill = fill

    workbook.save(filename=output_file)

if __name__ == "__main__":
    main()




# 307 to 311 matched in third pass - problem 

# 355 matched in first pass - problem

# 395 to 398 matched in first pass based on phone numbers - problem

# 449, 450 matched in first pass based on phone numbers - problem

# first pass problems fixed

# 307-311 incorrect matched in pass 3 - problem